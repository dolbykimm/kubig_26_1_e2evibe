import io
import math
import os
import re

import pandas as pd
import streamlit as st
from dotenv import load_dotenv
from groq import Groq
import openpyxl

from streamlit_gsheets import GSheetsConnection

load_dotenv()

# ─────────────────────────────────────────────────────────────
# Groq client
# ─────────────────────────────────────────────────────────────
_groq_client = None


def get_groq_client() -> Groq:
    global _groq_client
    if _groq_client is None:
        api_key = os.getenv("GROQ_API_KEY")
        if not api_key:
            raise ValueError("GROQ_API_KEY가 .env 파일에 설정되지 않았습니다.")
        _groq_client = Groq(api_key=api_key)
    return _groq_client


@st.cache_data(ttl=300, show_spinner=False)
def fetch_llama_models() -> list[str]:
    """Groq에서 현재 활성 모델 목록을 가져와 'llama' 포함 모델만 반환한다. (5분 캐시)"""
    try:
        client = get_groq_client()
        models = client.models.list().data
        filtered = sorted(
            m.id for m in models if "llama" in m.id.lower()
        )
        return filtered if filtered else ["llama-3.3-70b-versatile"]
    except Exception:
        return ["llama-3.3-70b-versatile"]  # API 실패 시 폴백


def analyze_personality(학과: str, 학번: str, 이름: str, 자소서: str, model: str) -> str:
    """자소서 텍스트를 받아 외향/내향 성격 요약 문자열을 반환한다."""
    prompt = f"""다음은 지원자의 자기소개서입니다.
학과: {학과} | 학번: {학번} | 이름: {이름}

[자기소개서]
{자소서}

이 지원자의 성격을 **외향형** 또는 **내향형** 중 하나로 판단하세요.

[외향형 판정 기준 — 아래 중 하나라도 해당하면 외향형으로 판단하세요]
① 팀장·조장·회장·반장 등 리더 역할을 맡은 경험
② 발표·토론·MC·사회 등 말하는 역할을 자발적으로 맡은 경험
③ 동아리·학생회·팀 프로젝트에서 주도적으로 이끈 경험
④ 말을 잘한다, 발표를 잘한다, 친화력이 좋다는 구체적 서술

[주의]
- 자기소개서는 자기 포장 글이므로 단순히 "활발하다", "사교적이다"라는 자기 주장만 있으면 신뢰도를 낮게 보세요.
- 위 ①~④ 중 구체적 근거가 하나도 없으면 내향형으로 판단하세요.
- 모호하면 내향형으로 판단하세요.

반드시 첫 줄에 "외향형" 또는 "내향형" 한 단어만 적고, 줄바꿈 후 2~3문장으로 근거를 작성하세요."""

    client = get_groq_client()
    response = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.3,
        max_tokens=256,
    )
    return response.choices[0].message.content.strip()


# ─────────────────────────────────────────────────────────────
# 성별 추정 (LLM 배치 방식)
# ─────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def estimate_genders_batch(names_tuple: tuple[str, ...], model: str) -> dict[str, str]:
    """
    이름 목록을 한 번의 LLM 호출로 성별 추정.
    캐시 키로 tuple 사용 (hashable).
    """
    names = list(names_tuple)
    if not names:
        return {}
    names_str = "\n".join(f"- {n}" for n in names)
    prompt = (
        "아래 한국 이름들의 성별을 추정해 반드시 '이름: 남' 또는 '이름: 여' 형식으로만 한 줄씩 답하라.\n"
        f"{names_str}"
    )
    try:
        client = get_groq_client()
        resp = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0,
            max_tokens=len(names) * 12 + 30,
        )
        result: dict[str, str] = {}
        for line in resp.choices[0].message.content.strip().split("\n"):
            for name in names:
                if name in line:
                    result[name] = "여" if "여" in line else "남"
                    break
        for name in names:
            result.setdefault(name, "미상")   # 누락된 이름 기본값
        return result
    except Exception:
        return {n: "미상" for n in names}


# ─────────────────────────────────────────────────────────────
# 참가자 명단 파일 파싱 (TXT / XLSX)
# ─────────────────────────────────────────────────────────────
def parse_roster_file(uploaded_file) -> pd.DataFrame | None:
    """
    TXT 또는 XLSX 참가자 명단 파일 → DataFrame 반환.
    - TXT : 줄마다 이름 하나 (공백 제거)
    - XLSX: resolve_columns 로 이름/학번/학과 자동 감지 후 전체 반환
    """
    fname = uploaded_file.name.lower()
    if fname.endswith(".txt"):
        text  = uploaded_file.read().decode("utf-8", errors="ignore")
        names = [l.strip() for l in text.splitlines() if l.strip()]
        return pd.DataFrame({"이름": names}) if names else None

    # XLSX / XLS
    df = pd.read_excel(uploaded_file)
    if df.empty:
        return None

    # 열 이름이 정수(openpyxl raw)면 첫 행을 헤더로 올림
    if all(isinstance(c, int) for c in df.columns):
        df.columns = df.iloc[0].astype(str)
        df = df.iloc[1:].reset_index(drop=True)

    # 이름 컬럼 감지: resolve_columns 우선, 없으면 _is_korean_name 점수 기반
    col_map  = resolve_columns(list(df.columns))
    name_col = col_map.get("이름")
    if name_col is None:
        best_col, best_cnt = None, 0
        for col in df.columns:
            cnt = df[col].dropna().astype(str).apply(_is_korean_name).sum()
            if cnt > best_cnt:
                best_cnt, best_col = cnt, col
        name_col = best_col

    if name_col is not None and name_col != "이름":
        df = df.rename(columns={name_col: "이름"})

    return df


# ─────────────────────────────────────────────────────────────
# 자리배치 알고리즘
# ─────────────────────────────────────────────────────────────
def assign_seats(
    df: pd.DataFrame,
    num_people: int,
    personality_mode: str,
    student_id_policy: str,
    model: str = "llama-3.3-70b-versatile",
) -> pd.DataFrame:
    """
    규칙
    ① E/I 는 테이블마다 무조건 고루 혼합 (round-robin offset)
    ② 성별도 가능한 한 혼합 (남·여 교차 정렬)
    ③ personality_mode
       - 혼합 배치 : 성별 교차 정렬 후 E/I 분산
       - 유사 배치 : 학과 기준으로 정렬 (같은 학과끼리 가깝게)
       - 무작위    : 무작위 셔플 후 E/I 분산
    ④ student_id_policy
       - 학번 무관          : 학번 무시
       - 동일 학번 분리     : 같은 연도 인원이 다른 테이블로 분산
       - 동일 학번 우선 배치: 같은 연도 인원이 같은 테이블에 집중

    Returns
    -------
    원본 컬럼 + ['성별', 'EI', '학번_연도', '테이블_번호'] 추가된 DataFrame
    (테이블_번호 오름차순 정렬)
    """
    work = df.copy().reset_index(drop=True)
    work["EI"] = work["성격 판정"].apply(lambda x: "E" if "외향" in str(x) else "I")

    # LLM 배치로 성별 추정 (한 번의 API 호출)
    unique_names   = tuple(work["이름"].astype(str).unique())
    gender_map     = estimate_genders_batch(unique_names, model)
    work["성별"]   = work["이름"].astype(str).map(gender_map).fillna("미상")

    work["학번_연도"] = work["학번"].astype(str).str[2:4]

    n          = len(work)
    num_tables = math.ceil(n / num_people)

    # ── 그룹별 정렬 순서 결정 ──────────────────────────────────
    def ordered_indices(mask: pd.Series) -> list[int]:
        sub = work[mask].copy()

        # 1) 학번 정책 선적용
        if student_id_policy == "동일 학번 분리":
            # 연도 오름차순 → round-robin 이 다른 테이블로 퍼뜨림
            sub = sub.sort_values("학번_연도")
        elif student_id_policy == "동일 학번 우선 배치":
            # 연도 내림차순 → 연속 할당으로 같은 테이블에 몰림
            sub = sub.sort_values("학번_연도", ascending=False)

        # 2) 성격 모드 적용
        if personality_mode == "무작위":
            return sub.sample(frac=1).index.tolist()

        if personality_mode == "유사 배치 (비슷한 성격끼리)":
            # 학과 기준 → 같은 학과끼리 인접
            sub = sub.sort_values(["학번_연도", "학과"])
            return sub.index.tolist()

        # 혼합 배치 (default): 남·여 교차 정렬
        males   = sub[sub["성별"] == "남"].index.tolist()
        females = sub[sub["성별"] == "여"].index.tolist()
        others  = sub[sub["성별"] == "미상"].index.tolist()
        merged  = []
        while males or females:
            if males:   merged.append(males.pop(0))
            if females: merged.append(females.pop(0))
        return merged + others

    e_idxs = ordered_indices(work["EI"] == "E")
    i_idxs = ordered_indices(work["EI"] == "I")

    # ── 테이블 할당 (E/I 교대 합산 후 단일 round-robin) ──────
    # E와 I를 교대로 합치면 자연스럽게 각 테이블에 E/I가 섞이고,
    # 단일 round-robin이므로 어떤 테이블도 ceil(n/num_tables) ≤ num_people 를 절대 초과하지 않음
    tables: list[list[int]] = [[] for _ in range(num_tables)]

    merged_idxs: list[int] = []
    for e, i in zip(e_idxs, i_idxs):
        merged_idxs.append(e)
        merged_idxs.append(i)
    merged_idxs.extend(e_idxs[len(i_idxs):])
    merged_idxs.extend(i_idxs[len(e_idxs):])

    # 늦참자를 큐 끝에 몰리지 않게 균등 삽입
    late_flag = work.get("늦참자", pd.Series(False, index=work.index)).fillna(False).astype(bool)
    late_set  = set(work[late_flag].index.tolist())
    if late_set:
        regular = [idx for idx in merged_idxs if idx not in late_set]
        lates   = [idx for idx in merged_idxs if idx in late_set]
        n_r, n_l = len(regular), len(lates)
        # step: 정규 인원 사이에 등간격으로 삽입
        step = (n_r + 1) / (n_l + 1)
        for i, li in enumerate(lates):
            pos = int(step * (i + 1)) + i
            regular.insert(min(pos, len(regular)), li)
        merged_idxs = regular

    for pos, idx in enumerate(merged_idxs):
        tables[pos % num_tables].append(idx)

    work["테이블_번호"] = 0
    for t_num, members in enumerate(tables):
        for idx in members:
            work.at[idx, "테이블_번호"] = t_num + 1

    return work.sort_values("테이블_번호").reset_index(drop=True)


def extract_interview_score(text: str) -> str:
    """면접 통합데이터에서 점수 숫자를 추출해 반환. 못 찾으면 빈 문자열.

    점수는 한 자리 정수(음수 포함)가 대부분이므로
    [총점]/[합계]/[점수] 레이블 뒤에 오는 -?\\d+ 패턴을 우선 찾는다.
    레이블이 없으면 빈 문자열 반환.
    """
    s = str(text)
    m = re.search(
        r"\[(총점|합계|점수|평균|score|total)\]\s*(-?[0-9]+(?:\.[0-9]+)?)",
        s, re.IGNORECASE
    )
    return m.group(2) if m else ""


# ─────────────────────────────────────────────────────────────
# 엑셀 직렬화
# ─────────────────────────────────────────────────────────────
def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # 시트1: 테이블별 전체 목록
        df.to_excel(writer, index=False, sheet_name="자리배치_전체")

        # 시트2: 테이블별 요약 (테이블_번호 × 이름·EI·성별)
        summary_rows = []
        for t_num, grp in df.groupby("테이블_번호"):
            members = ", ".join(
                f"{r['이름']}({r['EI']}/{r['성별']})"
                for _, r in grp.iterrows()
            )
            e_cnt = (grp["EI"] == "E").sum()
            i_cnt = (grp["EI"] == "I").sum()
            summary_rows.append({
                "테이블": t_num,
                "인원": len(grp),
                "E수": e_cnt,
                "I수": i_cnt,
                "구성": members,
            })
        pd.DataFrame(summary_rows).to_excel(writer, index=False, sheet_name="테이블_요약")

    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# 2단계: 면접표 병합 헬퍼
# ─────────────────────────────────────────────────────────────

_SHEET_DETAIL_LIMIT = 10   # 이 개수 이상이면 상세 샘플 생략
_SAMPLE_ROWS        = 3    # 시트당 샘플 행 수
_MAX_COL_WIDTH      = 20   # 셀 값 최대 표시 길이
_MAX_COLS_PER_SHEET = 8    # 시트당 최대 표시 열 수


def build_sheets_summary(sheets: dict[str, pd.DataFrame]) -> str:
    """
    토큰 절약형 시트 구조 요약.
    - 비어 있는 열 제외, 핵심 열만 최대 _MAX_COLS_PER_SHEET 개
    - 셀 값은 _MAX_COL_WIDTH 자로 절단
    - 시트 수 >= _SHEET_DETAIL_LIMIT 이면 시트 이름 목록만 전송
    """
    sheet_names = list(sheets.keys())

    # 시트가 너무 많으면 이름 목록만
    if len(sheets) >= _SHEET_DETAIL_LIMIT:
        return f"시트 목록({len(sheets)}개): {sheet_names}"

    parts = []
    for name, df in sheets.items():
        # 값이 하나라도 있는 열만 추림
        nonempty_cols = [c for c in df.columns if df[c].notna().any()]
        cols = nonempty_cols[:_MAX_COLS_PER_SHEET]

        # 샘플 행: 지정 열만, 셀 값 절단
        sample_df = df[cols].head(_SAMPLE_ROWS).astype(str)
        sample_df = sample_df.applymap(lambda v: v[:_MAX_COL_WIDTH])

        rows_txt = "; ".join(
            " | ".join(f"{c}={sample_df.at[i,c]}" for c in cols)
            for i in sample_df.index
        )
        parts.append(f"[{name}] 열:{cols} / 샘플:{rows_txt}")

    return "\n".join(parts)


def infer_sheet_structure(summary: str, model: str) -> str:
    """Groq LLM에게 시트 분류 기준과 코멘트 추출 방법을 추론하게 한다."""
    prompt = (
        "면접표 엑셀 시트 구조:\n"
        f"{summary}\n\n"
        "질문: 시트 분류 기준(면접관별/날짜별/조별 등)과 코멘트 열 위치를 한국어로 간결히 답하라."
    )

    client = get_groq_client()
    resp = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.2,
        max_tokens=256,
    )
    return resp.choices[0].message.content.strip()



def _find_header_row(df_raw: pd.DataFrame, max_scan: int = 10) -> int:
    """
    상위 max_scan 행을 순회해 ROLE_KEYWORDS 키워드를 가장 많이 포함한 행 인덱스를 반환한다.
    해당 행이 진짜 컬럼명 행으로 간주된다.
    """
    all_keywords = [kw for kws in ROLE_KEYWORDS.values() for kw in kws]
    best_row, best_count = 0, 0
    for i in range(min(max_scan, len(df_raw))):
        row_vals = df_raw.iloc[i].astype(str).str.lower()
        count = sum(any(kw in cell for kw in all_keywords) for cell in row_vals)
        if count > best_count:
            best_count, best_row = count, i
    return best_row


def _trim_tail(df: pd.DataFrame) -> pd.DataFrame:
    """
    위에서부터 순회해 '모든 열이 빈칸(NaN 또는 공백)'인 첫 행을 찾아
    그 행부터 아래를 모두 제거한다.
    → 데이터 블록 직후의 빈 행이 평가기준 표 등 푸터와의 경계선이 된다.
    """
    for i in range(len(df)):
        if df.iloc[i].apply(lambda v: pd.isna(v) or str(v).strip() == "").all():
            return df.iloc[:i]
    return df


# ─── 지원자 이름 감지 상수 ──────────────────────────────────────
_NAME_PREFIX_RE = re.compile(r"^([가-힣]{2,4})")  # 이름 파싱용 (prefix)
# 한국 성씨 목록 — 이름은 반드시 성씨로 시작한다는 점을 활용해 "활발함" 같은 코멘트 단어를 걸러냄
_KR_SURNAMES    = frozenset(
    "김이박최정강조윤장임한오서신권황안송류전홍고문손양배백허유남심노하곽성차주우구나민"
    "변엄원천방공현함변여추도소석길승라모봉표망제경은편심봉"
)
# 이름 열 헤더 레이블: 이 값만 있는 행은 data_start 후보에서 제외
_HEADER_NAMES   = frozenset({"이름", "성명", "성함", "지원자", "대상자"})


def _is_korean_name(v: str) -> bool:
    """
    셀 값이 한국 이름처럼 보이면 True.
    조건: ① 한글 2~4자로 시작하고 ② 첫 글자가 한국 성씨 목록에 있을 것.
    → "활발함"(활: 非성씨), "논리적"(논: 非성씨) 같은 코멘트 단어를 걸러냄.
    뒤에 "(컴공)", " 20240001" 등이 붙어있어도 OK.
    """
    s = v.strip()
    if s in _HEADER_NAMES or s in ("nan", "NaN", "None", ""):
        return False
    m = _NAME_PREFIX_RE.match(s)
    if not m:
        return False
    name_part = m.group(1)
    return len(name_part) >= 2 and name_part[0] in _KR_SURNAMES


def _find_name_col_idx(df_raw: pd.DataFrame) -> int | None:
    best_col, best_count = None, 0
    for ci in range(len(df_raw.columns)):
        # 이름이 1명만 있는 테스트 시트라도 통과하도록 >= 1로 수정
        cnt = sum(
            1 for v in df_raw.iloc[:, ci].astype(str)
            if _is_korean_name(v) and len(str(v).strip()) < 10
        )
        if cnt > best_count:
            best_count, best_col = cnt, ci
    return best_col if best_count >= 1 else None


def _parse_name_cell(raw: str) -> tuple[str, str, str]:
    """
    이름 셀에서 (순수이름, 학과힌트, 학번힌트)를 분리한다.

    예)  "박지성(컴공)"       → ("박지성", "컴공", "")
         "박지성 20240001"    → ("박지성", "",     "20240001")
         "박지성/통계 2024"   → ("박지성", "통계", "2024")
         "박지성"             → ("박지성", "",     "")
    """
    raw = raw.strip()
    m = _NAME_PREFIX_RE.match(raw)
    if not m:
        return raw, "", ""
    name = m.group(0)
    rest = raw[len(name):].strip().strip("()（）[][]〔〕/|- ")
    # 숫자 연속열(학번)
    id_m     = re.search(r"\d{4,}", rest)
    extra_id = id_m.group(0) if id_m else ""
    # 나머지 한글 부분(학과 힌트)
    extra_dept = re.sub(r"\d+", "", rest).strip("()（）[][]〔〕/|- ").strip()
    return name, extra_dept, extra_id


def extract_all_comments(sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    """
    상태 머신 방식으로 면접 통합데이터를 추출한다.

    1. 헤더 동적 탐색: '지원자'/'이름' AND '의견'/'비고' 키워드가 같은 행에 존재하는 행을 헤더로 특정
    2. current_name 상태 유지: 이름 열에 한국어 이름이 나오면 갱신
    3. 의견/비고 열 텍스트를 누적 (빈칸 아닌 모든 셀)
    4. 전체 병합 메모 행 처리: 열 헤더 무관하게 빈칸이 아닌 텍스트 → 현재 추적 지원자에 귀속
    5. '평가기준'/'평가 기준'/'주의사항' 발견 시 즉시 파싱 중단
    """
    _EMPTY = {"", "nan", "NaN", "None", "none"}
    _FOOTER_WORDS = ("평가기준", "평가 기준", "주의사항")
    _NAME_KW  = ("지원자", "이름", "성명", "성함", "대상자")
    _OPINE_KW = ("의견", "비고", "코멘트", "comment", "특이사항", "추가의견")

    def _cell(v) -> str:
        s = str(v).strip()
        return "" if s.lower() in _EMPTY else s

    def _row_has_footer(row_cells: list[str]) -> bool:
        joined = " ".join(row_cells)
        return any(w in joined for w in _FOOTER_WORDS)

    def _find_header_row(df_raw: pd.DataFrame) -> int | None:
        """'이름계열' 키워드 AND '의견계열' 키워드가 동시에 있는 첫 행 인덱스를 반환."""
        for ri in range(min(30, len(df_raw))):
            cells = [_cell(df_raw.iloc[ri, ci]) for ci in range(len(df_raw.columns))]
            cells_lower = [c.lower() for c in cells]
            has_name  = any(any(kw in c for kw in _NAME_KW)  for c in cells_lower)
            has_opine = any(any(kw in c for kw in _OPINE_KW) for c in cells_lower)
            if has_name and has_opine:
                return ri
        return None

    def _find_name_col_from_header(header_cells: list[str]) -> int | None:
        """헤더 행에서 이름 열 인덱스 반환."""
        for ci, c in enumerate(header_cells):
            if any(kw in c.lower() for kw in _NAME_KW):
                return ci
        return None

    def _find_opine_cols(header_cells: list[str]) -> list[int]:
        """헤더 행에서 의견/비고 열 인덱스 목록 반환."""
        return [ci for ci, c in enumerate(header_cells)
                if any(kw in c.lower() for kw in _OPINE_KW)]

    all_rows: list[dict] = []

    for sheet_name, df_raw in sheets.items():
        if df_raw.empty:
            continue

        # ── 1. 헤더 행 탐색 ──────────────────────────────────────────
        header_ri = _find_header_row(df_raw)
        if header_ri is None:
            # 의견 키워드 없어도 이름 열은 있을 수 있으므로 fallback: 이름 열만으로 탐색
            name_ci_fallback = _find_name_col_idx(df_raw)
            if name_ci_fallback is None:
                continue
            # 헤더는 이름이 처음 나오는 행 바로 위
            for ri in range(len(df_raw)):
                if _is_korean_name(_cell(df_raw.iloc[ri, name_ci_fallback])):
                    header_ri = max(0, ri - 1)
                    break
            if header_ri is None:
                continue

        header_cells = [_cell(df_raw.iloc[header_ri, ci]) for ci in range(len(df_raw.columns))]
        name_ci  = _find_name_col_from_header(header_cells)
        opine_cis = _find_opine_cols(header_cells)

        # 이름 열을 헤더에서 못 찾으면 전체 시트에서 재탐색
        if name_ci is None:
            name_ci = _find_name_col_idx(df_raw)
        if name_ci is None:
            continue

        # ── 2. 상태 머신: 헤더 다음 행부터 순회 ─────────────────────
        current_name     = ""
        current_raw_name = ""
        current_id       = ""
        current_dept     = ""
        current_phone    = ""
        current_parts: list[str] = []

        # id/학과/전화 열 인덱스 (헤더에서)
        id_ci_list    = [ci for ci, c in enumerate(header_cells) if "학번" in c or "번호" in c and "전화" not in c]
        dept_ci_list  = [ci for ci, c in enumerate(header_cells) if "학과" in c or "부서" in c or "전공" in c]
        phone_ci_list = [ci for ci, c in enumerate(header_cells) if "전화" in c or "연락" in c]
        id_ci    = id_ci_list[0]    if id_ci_list    else None
        dept_ci   = dept_ci_list[0]  if dept_ci_list  else None
        phone_ci  = phone_ci_list[0] if phone_ci_list else None
        score_cis = [ci for ci, c in enumerate(header_cells)
                     if any(kw in c for kw in ("총점", "합계", "점수"))]

        def _flush():
            """현재 지원자 데이터를 all_rows 에 저장."""
            if current_name and current_parts:
                all_rows.append({
                    "학번":        current_id,
                    "이름":        current_name,
                    "원본이름":    current_raw_name,
                    "학과":        current_dept,
                    "전화번호":    current_phone,
                    "면접_통합데이터": "\n".join(current_parts),
                })

        stop = False
        for ri in range(header_ri + 1, len(df_raw)):
            row_cells = [_cell(df_raw.iloc[ri, ci]) for ci in range(len(df_raw.columns))]

            # 꼬리 감지 → 즉시 중단
            if _row_has_footer(row_cells):
                stop = True
                break

            # 완전 빈 행 → 건너뜀
            if not any(row_cells):
                continue

            # 이름 열 확인
            name_val = row_cells[name_ci] if name_ci < len(row_cells) else ""
            if _is_korean_name(name_val):
                # 새 지원자 시작 → 이전 지원자 저장
                _flush()
                parsed_name, xdept, xid = _parse_name_cell(name_val)
                current_name     = parsed_name
                current_raw_name = name_val
                current_id       = (row_cells[id_ci] if id_ci is not None and id_ci < len(row_cells) else "") or xid
                current_dept     = (row_cells[dept_ci] if dept_ci is not None and dept_ci < len(row_cells) else "") or xdept
                current_phone    = row_cells[phone_ci] if phone_ci is not None and phone_ci < len(row_cells) else ""
                current_parts    = []

            if not current_name:
                # 아직 지원자 탐색 전 → 건너뜀
                continue

            # ── 의견/비고 열 텍스트 수집 ──────────────────────────────
            if opine_cis:
                for oci in opine_cis:
                    if oci < len(row_cells) and row_cells[oci]:
                        col_label = header_cells[oci] if oci < len(header_cells) else f"col{oci}"
                        current_parts.append(f"[{col_label}] {row_cells[oci]}")
                # 총점 계열은 의견 열 유무와 무관하게 항상 수집
                for sci in score_cis:
                    if sci not in opine_cis and sci < len(row_cells) and row_cells[sci]:
                        col_label = header_cells[sci] if sci < len(header_cells) else f"col{sci}"
                        current_parts.append(f"[{col_label}] {row_cells[sci]}")
            else:
                # 의견 열이 없으면: 이름/id/학과/전화 외 모든 비어있지 않은 셀 수집
                skip_cis = {c for c in [name_ci, id_ci, dept_ci, phone_ci] if c is not None}
                for ci, val in enumerate(row_cells):
                    if ci not in skip_cis and val:
                        col_label = header_cells[ci] if ci < len(header_cells) else f"col{ci}"
                        current_parts.append(f"[{col_label}] {val}")

            # ── 전체 병합 메모 행 처리 ────────────────────────────────
            # 이름 열이 비어있고 의견 열도 모두 비어있지만 다른 셀에 텍스트 있으면 → 병합 메모
            name_empty = not name_val
            opine_empty = not any(row_cells[oci] for oci in opine_cis if oci < len(row_cells)) if opine_cis else True
            if name_empty and opine_empty:
                skip_cis = {c for c in [name_ci, id_ci, dept_ci, phone_ci] if c is not None}
                memo_vals = [val for ci, val in enumerate(row_cells) if ci not in skip_cis and val]
                if memo_vals:
                    current_parts.append("[메모] " + " / ".join(memo_vals))

        # 마지막 지원자 저장
        if not stop or current_name:
            _flush()

    if not all_rows:
        return pd.DataFrame(columns=["학번", "이름", "원본이름", "학과", "전화번호", "면접_통합데이터"])

    df_res = pd.DataFrame(all_rows)
    df_res["이름"] = df_res["이름"].astype(str).str.strip()
    df_res["학번"] = df_res["학번"].astype(str).str.strip()

    # cross-sheet 동일인 합산: 학번이 있는 레코드만 groupby로 합침.
    # 학번이 없는 레코드는 각각 독립 유지 (동명이인 두 명을 한 명으로 합치는 버그 방지).
    _AGG = {
        "원본이름":        "first",
        "학과":            "first",
        "전화번호":        "first",
        "면접_통합데이터": lambda x: "\n---\n".join(filter(None, x)),
    }
    df_with_id = df_res[df_res["학번"].str.len() > 0]
    df_no_id   = df_res[df_res["학번"].str.len() == 0]

    merged_id = (
        df_with_id.groupby(["학번", "이름"], as_index=False).agg(_AGG)
        if not df_with_id.empty else df_with_id
    )
    return pd.concat([merged_id, df_no_id], ignore_index=True)


def smart_merge(
    df_base: pd.DataFrame,
    df_comments: pd.DataFrame,
) -> tuple[pd.DataFrame, list[dict]]:
    """
    병합 규칙:
    1. 학번 완전 일치 → 자동 확정
    2. 이름 일치 + 후보 1명 → 자동 확정
    3. 동명이인(후보 2명+) → 전원 수동 처리
    """
    df_merged = df_base.copy()
    df_merged["면접_통합데이터"] = None
    ambiguous: list[dict] = []

    for c_idx, c_row in df_comments.iterrows():
        c_이름     = str(c_row["이름"]).strip()
        c_원본이름 = str(c_row.get("원본이름", c_이름)).strip()
        c_학번     = str(c_row.get("학번", "")).strip()
        c_학과     = str(c_row.get("학과", "")).strip()
        comment    = str(c_row["면접_통합데이터"])

        name_mask  = df_base["이름"].astype(str).str.strip() == c_이름
        candidates = df_base[name_mask]

        if candidates.empty:
            continue

        # 학번 완전 일치 → 자동 확정
        if c_학번:
            exact = candidates[candidates["학번"].astype(str).str.strip() == c_학번]
            if len(exact) == 1:
                df_merged.at[exact.index[0], "면접_통합데이터"] = comment
                continue

        # 후보 1명 → 자동 확정
        if len(candidates) == 1:
            df_merged.at[candidates.index[0], "면접_통합데이터"] = comment
            continue

        # 동명이인 → 수동 처리
        ambiguous.append({
            "comment_idx":  c_idx,
            "name":         c_이름,
            "원본이름":     c_원본이름,
            "학과":         c_학과,
            "학번":         c_학번,
            "comment_text": comment,
            "candidates": [
                {
                    "base_idx": idx,
                    "label": (
                        f"{row['이름']} "
                        f"({row.get('학과', '학과?')} / "
                        f"학번: {row.get('학번', '?')})"
                    ),
                }
                for idx, row in candidates.iterrows()
            ],
        })

    return df_merged, ambiguous


def reanalyze_final(
    이름: str,
    성격_판정: str,
    근거: str,
    면접_통합데이터: str,
    model: str,
) -> str:
    """자소서 기반 판정 + 면접 통합 데이터를 종합해 최종 성격을 판단한다."""
    prompt = (
        f"지원자: {이름}\n\n"
        f"[자소서 기반 1차 판정 (참고용)]\n{성격_판정} — {근거}\n\n"
        f"[면접 현장 데이터]\n{str(면접_통합데이터).strip()}\n\n"
        "위 데이터를 바탕으로 다음 기준에 따라 최종 성격을 판단하세요.\n\n"
        "[외향형 판정 조건 — 아래 중 하나라도 해당하면 외향형으로 판단하세요]\n"
        "  ① 면접관이 '말을 잘한다', '발표력이 좋다', '활발하다', '말이 많다', '적극적' 등 언어 표현력·에너지를 직접 언급\n"
        "  ② 면접관이 팀장·조장·리더 역할을 언급하거나 주도성을 긍정적으로 평가\n"
        "  ③ 면접 점수가 높음 (점수 자체가 높으면 면접장에서 자신을 잘 표현했다는 신호이므로 외향형 가능성이 높음)\n"
        "  ④ 자소서에서 팀장·발표·MC 등 리더/스피커 역할 경험이 구체적으로 확인됨\n\n"
        "[가중치]\n"
        "• 면접 코멘트 > 면접 점수 > 자소서 판정 순으로 우선합니다.\n"
        "• 면접 데이터와 자소서 판정이 충돌하면 면접 데이터를 따르세요.\n"
        "• 위 ①~④ 중 근거가 하나도 없을 때만 내향형으로 판단하세요.\n\n"
        "다음 형식으로 정확히 3줄만 출력하세요:\n"
        "1줄: '외향형' 또는 '내향형' 한 단어\n"
        "2줄: 이 사람을 대표하는 성격 키워드를 쉼표로 나열 (예: 발표경험 多, 적극적, 리더십) — 20자 이내\n"
        "3줄: 판정 근거 한 문장"
    )

    client = get_groq_client()
    resp = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": prompt}],
        temperature=0.3,
        max_tokens=200,
    )
    return resp.choices[0].message.content.strip()


def to_final_excel_bytes(df: pd.DataFrame) -> bytes:
    # 원하는 열 순서: 성격판정 → 키워드 → 평균점수 → 근거 → 면접데이터 → 나머지
    preferred = ["최종_성격_판정", "성격_키워드", "면접_평균점수", "최종_근거", "면접_통합데이터"]
    # 실제로 df에 존재하는 열만 preferred 순서로, 없는 열은 끝에 추가
    ordered = [c for c in preferred if c in df.columns]
    rest    = [c for c in df.columns if c not in ordered]
    df_out  = df[rest + ordered]

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False, sheet_name="최종_분석")
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────
# Google Sheets 공유
# ─────────────────────────────────────────────────────────────

# session_state 키 → 워크시트 이름 매핑
GS_TARGETS: list[tuple[str, str, str]] = [
    ("자소서 분석 결과", "자소서분석", "df_personality"),
    ("최종 병합 분석",   "최종분석",   "df_merged"),
    ("자리배치 결과",    "자리배치",   "df_seated"),
]


@st.cache_resource(show_spinner=False)
def get_gsheets_conn():
    """GSheetsConnection 캐시 싱글턴. secrets.toml 미설정 시 None 반환."""
    try:
        return st.connection("gsheets", type=GSheetsConnection)
    except Exception:
        return None


def save_to_gsheets(df: pd.DataFrame, worksheet: str) -> None:
    conn = get_gsheets_conn()
    if conn is None:
        raise RuntimeError("Google Sheets 연결 미설정 — .streamlit/secrets.toml을 확인하세요.")
    # 모든 열을 str 변환해 Sheets API 직렬화 오류 방지
    conn.update(worksheet=worksheet, data=df.astype(str))


def load_from_gsheets(worksheet: str) -> pd.DataFrame:
    conn = get_gsheets_conn()
    if conn is None:
        raise RuntimeError("Google Sheets 연결 미설정 — .streamlit/secrets.toml을 확인하세요.")
    return conn.read(worksheet=worksheet, ttl=0)  # ttl=0: 항상 최신 데이터


# ─────────────────────────────────────────────────────────────
# 유연한 열 이름 매핑
# ─────────────────────────────────────────────────────────────
# 각 의미 역할(role)에 대해 엑셀에서 쓰일 법한 동의어를 모두 나열한다.
# 실제 열 이름과 키워드 사이의 '포함 관계'로 점수를 매겨 가장 근접한 열을 선택한다.
ROLE_KEYWORDS: dict[str, list[str]] = {
    "학과":   ["학과", "전공", "계열", "학부", "부서", "소속", "department", "major"],
    "학번":   ["학번", "학생번호", "번호", "student", "id", "학생id", "학생 번호"],
    "이름":   ["이름", "성명", "성함", "name", "지원자", "대상자"],
    "전화번호": ["전화", "연락처", "핸드폰", "휴대폰", "전화번호", "phone", "tel", "mobile"],
}


def _score(col: str, keyword: str) -> int:
    """열 이름과 키워드 사이의 유사도 점수."""
    col_str = str(col) if col is not None else ""
    c = col_str.strip().lower().replace(" ", "")
    k = keyword.strip().lower().replace(" ", "")
    if c == k:
        return 3          # 완전 일치
    if k in c or c in k:
        return 2          # 부분 문자열 포함
    # 글자 단위 겹침 비율 (짧은 쪽 기준)
    common = sum(1 for ch in k if ch in c)
    if len(k) and common / len(k) >= 0.6:
        return 1          # 60 % 이상 글자 겹침
    return 0


def resolve_columns(columns: list[str]) -> dict[str, str | None]:
    """
    엑셀 열 목록 → {role: 실제열이름} 매핑 반환.
    매칭 안 되면 None.
    """
    mapping: dict[str, str | None] = {}
    for role, keywords in ROLE_KEYWORDS.items():
        best_col, best = None, 0
        for col in columns:
            for kw in keywords:
                s = _score(col, kw)
                if s > best:
                    best, best_col = s, col
        mapping[role] = best_col if best > 0 else None
    return mapping


# ─────────────────────────────────────────────────────────────
# Streamlit UI
# ─────────────────────────────────────────────────────────────
st.set_page_config(page_title="OT 자리배치", layout="wide", page_icon="🪑")
st.title("🪑 OT 자리배치 앱")

# ── 사이드바 ─────────────────────────────────────────────────
with st.sidebar:
    # ── 모델 설정 ────────────────────────────────────────────
    st.header("모델 설정")
    with st.spinner("Groq 모델 목록 로딩 중..."):
        llama_models = fetch_llama_models()
    selected_model: str = st.selectbox(
        "사용할 LLM 모델",
        options=llama_models,
        index=0,
        help="Groq에서 현재 활성화된 llama 계열 모델 목록입니다. (5분마다 자동 갱신)",
    )
    st.caption(f"선택된 모델: `{selected_model}`")

    st.divider()
    st.subheader("📋 진행 현황")
    step1_done = "df_resume_raw"  in st.session_state
    step2_done = "df_personality" in st.session_state
    step3_done = "df_ei_adjusted" in st.session_state
    step4_done = "df_seated"      in st.session_state
    st.markdown(
        f"{'✅' if step1_done else '⬜'} STEP 1 · 파일 업로드\n\n"
        f"{'✅' if step2_done else '⬜'} STEP 2 · 분석 완료\n\n"
        f"{'✅' if step3_done else '⬜'} STEP 3 · 자리배치"
    )

    # ── Google Sheets 공유 ───────────────────────────────────
    st.divider()
    st.header("Google Sheets 공유")

    gsheets_ok = get_gsheets_conn() is not None
    if not gsheets_ok:
        st.warning("`secrets.toml` 미설정 — Google Sheets 기능 비활성화")

    for label, ws, ss_key in GS_TARGETS:
        st.markdown(f"**{label}**")
        c_save, c_load = st.columns(2)

        with c_save:
            if st.button(
                "저장",
                key=f"gs_save_{ws}",
                disabled=not (gsheets_ok and ss_key in st.session_state),
                use_container_width=True,
            ):
                try:
                    save_to_gsheets(st.session_state[ss_key], ws)
                    st.success("저장 완료")
                except Exception as e:
                    st.error(str(e))

        with c_load:
            if st.button(
                "불러오기",
                key=f"gs_load_{ws}",
                disabled=not gsheets_ok,
                use_container_width=True,
            ):
                try:
                    df_loaded = load_from_gsheets(ws)
                    st.session_state[ss_key] = df_loaded
                    st.success(f"{len(df_loaded)}행 로드")
                    st.rerun()
                except Exception as e:
                    st.error(str(e))


# ── 앱 소개 ──────────────────────────────────────────────────
with st.expander("💡 시작하기 전에 읽어주세요!", expanded=False):
    st.markdown("""
#### 이 앱은 대체 뭐 하는 건가요?
자기소개서와 면접 결과를 바탕으로 **OT 자리를 자동으로 배치**해 드려요.
외향·내향, 학번, 성별이 골고루 섞이도록 AI가 계산해요.

---

#### 사용 순서

| 단계 | 하는 일 |
|------|---------|
| **📁 STEP 1 · 준비** | 파일을 한 번에 다 올리고, 임원진·늦참자·취소자를 입력해요 |
| **🔍 STEP 2 · 분석** | 취소자를 제외한 명단을 확인한 뒤 분석 버튼을 딸깍딸깍 눌러요 |
| **🪑 STEP 3 · 자리배치** | 설정하고 버튼 누르면 끝! |

> 💡 면접표가 없어도 STEP 1 → STEP 2(자소서 분석만) → STEP 3 으로 배치 가능해요!

---

#### 자주 묻는 것들
- **임원진·기존부원은요?** → STEP 1에서 이름 직접 입력. 자동으로 외향으로 분류돼요.
- **늦참자가 있어요** → STEP 1에서 이름 입력. 나머지 인원과 골고루 섞어 배치돼요.
- **참석 취소자가 생겼어요** → STEP 1 취소자 칸에 이름 입력. 분석과 배치 모두에서 제외돼요.
- **Groq API 키는 왼쪽 사이드바** ⬅ 에서 확인하세요.
""")

tab1, tab2, tab3, tab4 = st.tabs([
    "📁  STEP 1 · 준비",
    "🔍  STEP 2 · 분석",
    "✏️  STEP 3 · 성격 조정",
    "🪑  STEP 4 · 자리배치",
])

# ── 1단계: 파일 업로드 & 수동 입력 ───────────────────────────
with tab1:
    st.header("📁 STEP 1 · 준비")
    st.caption("파일을 모두 올리고 추가 정보를 입력해 주세요. 다 됐으면 STEP 2로 이동하세요!")

    col_files, col_manual = st.columns([1, 1])

    with col_files:
        st.subheader("파일 업로드")

        _uploaded_resume = st.file_uploader(
            "📄 자기소개서 엑셀 (필수)",
            type=["xlsx", "xls"],
            key="file_resume",
            help="지원자의 자기소개서가 담긴 엑셀 파일을 올려주세요.",
        )
        if _uploaded_resume is not None:
            if st.session_state.get("_resume_file_name") != _uploaded_resume.name:
                _rb = _uploaded_resume.read()
                st.session_state["_resume_bytes"]     = _rb
                st.session_state["_resume_file_name"] = _uploaded_resume.name
                st.session_state["df_resume_raw"]     = pd.read_excel(io.BytesIO(_rb))
                for _k in ["df_personality", "df_merged", "df_comments", "ambiguous",
                            "sheet_inference", "df_seated"]:
                    st.session_state.pop(_k, None)
            st.success(f"자소서 파일 업로드 완료! ({len(st.session_state['df_resume_raw'])}명)")

        st.divider()

        _uploaded_interview = st.file_uploader(
            "📊 면접 점수표 엑셀 (선택)",
            type=["xlsx", "xls"],
            key="file_interview",
            help="면접 코멘트와 점수가 담긴 엑셀 파일이에요. 없으면 건너뛰세요.",
        )
        if _uploaded_interview is not None:
            if st.session_state.get("_interview_file_name") != _uploaded_interview.name:
                _raw = _uploaded_interview.read()
                _wb  = openpyxl.load_workbook(io.BytesIO(_raw), data_only=True)
                _sheets: dict[str, pd.DataFrame] = {}
                for _sname in _wb.sheetnames:
                    _ws   = _wb[_sname]
                    _rows = [[cell.value for cell in row] for row in _ws.iter_rows()]
                    _sheets[_sname] = pd.DataFrame(_rows) if _rows else pd.DataFrame()
                st.session_state["interview_sheets"]     = _sheets
                st.session_state["_interview_file_name"] = _uploaded_interview.name
                for _k in ["df_comments", "df_merged", "ambiguous", "sheet_inference", "df_seated"]:
                    st.session_state.pop(_k, None)
            _n_sheets = len(st.session_state.get("interview_sheets", {}))
            st.success(f"면접 파일 업로드 완료! (시트 {_n_sheets}개)")

        st.divider()

        _uploaded_roster = st.file_uploader(
            "📋 참가자 명단 (선택, txt 또는 엑셀)",
            type=["txt", "xlsx", "xls"],
            key="file_roster",
            help="참석자 이름 목록 파일이에요. 없으면 자소서 파일의 이름 목록을 쓸게요.",
        )
        if _uploaded_roster is not None:
            if st.session_state.get("_roster_file_name") != _uploaded_roster.name:
                _df_roster = parse_roster_file(_uploaded_roster)
                st.session_state["df_roster"]         = _df_roster
                st.session_state["_roster_file_name"] = _uploaded_roster.name
            _df_r = st.session_state.get("df_roster")
            if _df_r is not None and not _df_r.empty:
                st.success(f"참가자 명단 {len(_df_r)}명 인식 완료!")
                with st.expander("명단 미리보기"):
                    st.dataframe(_df_r, use_container_width=True)
            else:
                st.warning("명단을 인식하지 못했어요. 파일 형식을 확인해 주세요.")

    with col_manual:
        st.subheader("추가 인원 & 조정")

        st.markdown("**👑 임원진 / 기존 부원**")
        st.caption("한 줄에 한 명씩 입력하세요. 이름만 쓰면 외향형, 이름 뒤에 내향/내향형을 붙이면 내향형으로 처리돼요.")
        st.text_area(
            "임원진/기존부원 이름",
            placeholder="회장 홍길동\n부회장 김철수\n이영희",
            key="officer_input",
            label_visibility="collapsed",
            height=90,
        )

        st.divider()

        st.markdown("**⏰ 늦참자**")
        st.caption("나중에 도착하는 사람. 나머지 참가자와 골고루 섞어 배치돼요.")
        st.text_area(
            "늦참자 이름",
            placeholder="홍길동\n김철수",
            key="late_arrivals_input",
            label_visibility="collapsed",
            height=90,
        )

        st.divider()

        st.markdown("**❌ 취소자**")
        st.caption("참석하겠다고 했다가 취소한 사람. 분석과 자리배치 모두에서 제외돼요.")
        st.text_area(
            "취소자 이름",
            placeholder="홍길동\n김철수",
            key="cancellation_input",
            label_visibility="collapsed",
            height=90,
        )

    # ── 준비 현황 요약 ──────────────────────────────────────
    st.divider()
    _cancel_set = {
        n.strip()
        for n in st.session_state.get("cancellation_input", "").splitlines()
        if n.strip()
    }
    _checks = []
    if "df_resume_raw" in st.session_state:
        _checks.append(f"✅ 자소서 파일 — {len(st.session_state['df_resume_raw'])}명")
    else:
        _checks.append("⬜ 자소서 파일 **(필수)**")
    if "interview_sheets" in st.session_state:
        _checks.append(f"✅ 면접 파일 — 시트 {len(st.session_state['interview_sheets'])}개")
    else:
        _checks.append("⬜ 면접 파일 (선택)")
    if _cancel_set:
        _checks.append(f"❌ 취소자 {len(_cancel_set)}명 제외 예정")
    for _c in _checks:
        st.markdown(_c)
    if "df_resume_raw" in st.session_state:
        st.info("준비 완료! **STEP 2** 탭으로 이동해서 분석을 시작하세요.")


# ── 2단계: 분석 ───────────────────────────────────────────────
with tab2:
    st.header("🔍 STEP 2 · 분석")

    if "df_resume_raw" not in st.session_state:
        st.warning("먼저 **STEP 1**에서 자소서 파일을 올려주세요!")
    else:
        df_resume = st.session_state["df_resume_raw"]

        _cancel_set2 = {
            n.strip()
            for n in st.session_state.get("cancellation_input", "").splitlines()
            if n.strip()
        }

        # ── 열 이름 확인 ────────────────────────────────────
        st.subheader("열 이름 확인")
        st.caption("자동으로 찾아봤는데, 틀린 게 있으면 직접 바꿔주세요.")

        auto_map    = resolve_columns(list(df_resume.columns))
        col_options = [None] + list(df_resume.columns)
        role_labels = {
            "학과": "학과 / 전공",
            "학번": "학번 / 학생번호",
            "이름": "이름 / 성명",
        }

        confirmed_map: dict[str, str | None] = {}
        map_cols = st.columns(3)
        for _i, (_role, _label) in enumerate(role_labels.items()):
            _detected = auto_map[_role]
            _idx = col_options.index(_detected) if _detected in col_options else 0
            _chosen = map_cols[_i].selectbox(
                _label,
                options=col_options,
                index=_idx,
                format_func=lambda x: "— 선택 안 됨 —" if x is None else x,
                key=f"col_map_{_role}",
            )
            confirmed_map[_role] = _chosen

        info_cols      = {c for c in confirmed_map.values() if c is not None}
        remaining_cols = [c for c in df_resume.columns if c not in info_cols]

        st.markdown("**자기소개서 내용이 있는 칸을 골라주세요** (여러 개 선택 가능 — 선택한 순서대로 AI에 전달돼요)")
        essay_cols: list[str] = st.multiselect(
            label="자소서 항목 열",
            options=list(df_resume.columns),
            default=remaining_cols,
            key="essay_cols",
            label_visibility="collapsed",
        )

        # ── 참가자 명단 필터링 + 취소자 제외 ──────────────────
        df_resume_filtered = df_resume.copy()
        if confirmed_map.get("이름"):
            _name_col = confirmed_map["이름"]
            # ① 참가자 명단이 있으면 명단에 있는 사람만 남김
            _df_roster = st.session_state.get("df_roster")
            if _df_roster is not None and not _df_roster.empty and "이름" in _df_roster.columns:
                _roster_names = set(_df_roster["이름"].astype(str).str.strip())
                _before = len(df_resume_filtered)
                df_resume_filtered = df_resume_filtered[
                    df_resume_filtered[_name_col].astype(str).str.strip().isin(_roster_names)
                ].reset_index(drop=True)
                _roster_filtered = _before - len(df_resume_filtered)
                if _roster_filtered > 0:
                    st.info(f"참가자 명단 기준으로 필터링: {_roster_filtered}명 제외, 대상 {len(df_resume_filtered)}명")
            # ② 취소자 제외
            if _cancel_set2:
                _keep = ~df_resume_filtered[_name_col].astype(str).str.strip().isin(_cancel_set2)
                _removed = int((~_keep).sum())
                df_resume_filtered = df_resume_filtered[_keep].reset_index(drop=True)
                if _removed > 0:
                    st.info(f"취소자 {_removed}명이 제외됐어요. 분석 대상: **{len(df_resume_filtered)}명**")

        # ── 분석 대상 명단 미리보기 ──────────────────────────
        st.subheader("분석 대상 명단")
        st.dataframe(df_resume_filtered, use_container_width=True)

        unresolved = [lbl for role, lbl in role_labels.items() if confirmed_map[role] is None]

        st.divider()

        # ── ① 자소서 성격 분석 ──────────────────────────────
        st.subheader("① 자소서 성격 분석")

        if unresolved:
            st.warning(f"아직 연결 안 된 항목: **{', '.join(unresolved)}** — 위 드롭다운에서 골라주세요!")
        elif not essay_cols:
            st.warning("자기소개서 내용이 있는 칸을 하나 이상 골라야 해요!")
        else:
            if st.button("자소서 성격 분석 시작", type="primary", key="btn_analyze"):
                results = []
                progress = st.progress(0, text="분석 중...")
                total = len(df_resume_filtered)

                for _i, (_, row) in enumerate(df_resume_filtered.iterrows()):
                    combined_essay = "\n\n".join(
                        f"[{col}]\n{row[col]}"
                        for col in essay_cols
                        if pd.notna(row[col]) and str(row[col]).strip()
                    )
                    try:
                        summary = analyze_personality(
                            학과=str(row[confirmed_map["학과"]]),
                            학번=str(row[confirmed_map["학번"]]),
                            이름=str(row[confirmed_map["이름"]]),
                            자소서=combined_essay,
                            model=selected_model,
                        )
                        lines = summary.splitlines()
                        판정 = lines[0].strip() if lines else "알 수 없음"
                        근거 = " ".join(l.strip() for l in lines[1:] if l.strip())
                    except Exception as e:
                        판정, 근거 = "오류", str(e)

                    results.append({
                        "학과":    row[confirmed_map["학과"]],
                        "학번":    row[confirmed_map["학번"]],
                        "이름":    row[confirmed_map["이름"]],
                        "성격 판정": 판정,
                        "근거 요약": 근거,
                    })
                    progress.progress((_i + 1) / total, text=f"분석 중... ({_i + 1}/{total})")

                progress.empty()
                df_result = pd.DataFrame(results)
                st.session_state["df_personality"] = df_result
                st.success("분석 완료!")

            if "df_personality" in st.session_state:
                _df_p = st.session_state["df_personality"]
                st.dataframe(_df_p, use_container_width=True)

                # ── 성격 판정 직접 수정 ──────────────────────
                with st.expander("✏️ 성격 판정 직접 수정하기", expanded=False):
                    st.caption("AI 판정이 틀렸다면 여기서 바꿔주세요. 이름 옆 근거 요약도 참고하세요.")
                    _edited = False
                    _df_edit = _df_p.copy()
                    for _ei, _erow in _df_p.iterrows():
                        _ecols = st.columns([2, 2, 3])
                        _ecols[0].markdown(f"**{_erow['이름']}**")
                        _cur = "외향형" if "외향" in str(_erow["성격 판정"]) else "내향형"
                        _new_val = _ecols[1].radio(
                            "판정",
                            options=["외향형", "내향형"],
                            index=0 if _cur == "외향형" else 1,
                            horizontal=True,
                            key=f"edit_ei_{_ei}",
                            label_visibility="collapsed",
                        )
                        _ecols[2].caption(str(_erow.get("근거 요약", ""))[:60])
                        if _new_val != _erow["성격 판정"]:
                            _df_edit.at[_ei, "성격 판정"] = _new_val
                            _edited = True
                    if st.button("수정 적용", key="btn_apply_ei_edit"):
                        st.session_state["df_personality"] = _df_edit
                        st.success("성격 판정이 업데이트됐어요!")
                        st.rerun()

        # ── ② 면접표 병합 ────────────────────────────────────
        if "interview_sheets" in st.session_state:
            sheets = st.session_state["interview_sheets"]
            st.divider()
            st.subheader("② 면접표 병합")

            for sheet_name, df_s in sheets.items():
                with st.expander(f"시트 미리보기: {sheet_name}  ({len(df_s)}행 × {len(df_s.columns)}열)"):
                    st.dataframe(df_s, use_container_width=True)

            if st.button("AI로 시트 구조 분석", key="btn_infer"):
                with st.spinner("LLM이 시트 구조를 분석 중..."):
                    sheet_summary = build_sheets_summary(sheets)
                    inference     = infer_sheet_structure(sheet_summary, selected_model)
                st.session_state["sheet_inference"] = inference

            if "sheet_inference" in st.session_state:
                st.info(st.session_state["sheet_inference"])

            _has_personality = "df_personality" in st.session_state
            if not _has_personality:
                st.warning("먼저 위 ①에서 자소서 성격 분석을 완료해 주세요!")

            if st.button("면접표 병합하기", type="primary", disabled=not _has_personality, key="btn_merge"):
                with st.spinner("면접 코멘트 추출 중..."):
                    df_comments = extract_all_comments(sheets)

                df_base = st.session_state["df_personality"].copy()
                df_merged_new, ambiguous_new = smart_merge(df_base, df_comments)

                matched = df_merged_new["면접_통합데이터"].notna().sum()
                st.session_state["df_merged"]   = df_merged_new
                st.session_state["ambiguous"]   = ambiguous_new
                st.session_state["df_comments"] = df_comments

                if ambiguous_new:
                    st.warning(
                        f"자동 연결 완료 — {matched}/{len(df_merged_new)}명 | "
                        f"같은 이름이 **{len(ambiguous_new)}건** 있어요. 아래에서 직접 연결해 주세요! ↓"
                    )
                else:
                    st.success(f"완료! {matched} / {len(df_merged_new)}명 면접 코멘트 연결됐어요.")

            # ── 동명이인 수동 매칭 ───────────────────────────
            if st.session_state.get("ambiguous"):
                ambiguous_cases = st.session_state["ambiguous"]
                st.divider()
                st.subheader("같은 이름 직접 연결하기")
                st.caption("같은 이름이 여러 명이에요! 각 면접 기록이 누구 건지 직접 골라주세요.")

                from collections import defaultdict
                groups: dict[str, list[dict]] = defaultdict(list)
                for case in ambiguous_cases:
                    groups[case["name"]].append(case)

                for name, cases in groups.items():
                    st.markdown(f"### 동명이인: **{name}** ({len(cases)}건)")
                    candidates = cases[0]["candidates"]
                    cand_options = ["— 매칭 안 함 —"] + [c["label"] for c in candidates]

                    for case in cases:
                        raw  = case.get("원본이름", name)
                        dept = case.get("학과", "")
                        sid  = case.get("학번", "")
                        info_parts = [f"원본: **{raw}**"]
                        if dept: info_parts.append(f"학과: {dept}")
                        if sid:  info_parts.append(f"학번: {sid}")
                        info_str = " / ".join(info_parts)

                        preview = case["comment_text"][:150].replace("\n", " ")
                        if len(case["comment_text"]) > 150:
                            preview += "…"

                        with st.container(border=True):
                            st.markdown(info_str)
                            if not sid:
                                st.warning("학번 정보가 없어요 — 이름이나 코멘트 내용으로 구분해 주세요.")
                            st.caption(preview)
                            st.selectbox(
                                "→ 누구의 면접 기록인가요?",
                                options=cand_options,
                                key=f"manual_{case['comment_idx']}",
                            )

                    chosen_labels = [
                        st.session_state.get(f"manual_{c['comment_idx']}", "— 매칭 안 함 —")
                        for c in cases
                    ]
                    non_skip = [l for l in chosen_labels if l != "— 매칭 안 함 —"]
                    if len(non_skip) != len(set(non_skip)):
                        st.warning("같은 사람한테 두 개 이상 연결됐어요! 확인해 주세요.")
                    st.divider()

                if st.button("수동 매칭 확정", type="primary", key="btn_manual_confirm"):
                    all_chosen = [
                        st.session_state.get(f"manual_{case['comment_idx']}", "— 매칭 안 함 —")
                        for case in ambiguous_cases
                    ]
                    non_skip_all = [l for l in all_chosen if l != "— 매칭 안 함 —"]
                    if len(non_skip_all) != len(set(non_skip_all)):
                        st.error("같은 사람에게 면접 기록이 두 개 이상 연결됐어요! 수정 후 다시 확정해 주세요.")
                    else:
                        df_m = st.session_state["df_merged"].copy()
                        for case in ambiguous_cases:
                            chosen_label = st.session_state.get(
                                f"manual_{case['comment_idx']}", "— 매칭 안 함 —"
                            )
                            if chosen_label == "— 매칭 안 함 —":
                                continue
                            chosen_idx = next(
                                c["base_idx"]
                                for c in case["candidates"]
                                if c["label"] == chosen_label
                            )
                            df_m.at[chosen_idx, "면접_통합데이터"] = case["comment_text"]

                        st.session_state["df_merged"]  = df_m
                        st.session_state["ambiguous"]  = []
                        matched_final = df_m["면접_통합데이터"].notna().sum()
                        st.success(f"연결 완료! 최종 {matched_final}/{len(df_m)}명 연결됐어요.")
                        st.rerun()

            if "df_merged" in st.session_state:
                st.dataframe(st.session_state["df_merged"], use_container_width=True)

                st.divider()

                # ── ③ AI 재분석 ─────────────────────────────
                st.subheader("③ AI 재분석 (선택)")
                st.caption("자소서 분석 결과와 면접 코멘트를 합쳐서 AI가 다시 성격을 판단해요.")

                df_merged = st.session_state["df_merged"]
                has_comments = df_merged["면접_통합데이터"].notna().any()
                if not has_comments:
                    st.info("연결된 면접 코멘트가 없어서 재분석을 건너뛸게요.")

                if st.button("AI 재분석하기!", type="primary", disabled=not has_comments, key="btn_reanalyze"):
                    df_final = df_merged.copy()
                    df_final["최종_성격_판정"] = df_final["성격 판정"]
                    df_final["성격_키워드"]    = ""
                    df_final["면접_평균점수"]  = ""
                    df_final["최종_근거"]      = df_final["근거 요약"]

                    targets  = df_final[df_final["면접_통합데이터"].notna()]
                    progress = st.progress(0, text="재분석 중...")

                    for _i, (idx, row) in enumerate(targets.iterrows()):
                        raw_comment = str(row["면접_통합데이터"]).strip()
                        try:
                            result = reanalyze_final(
                                이름=str(row["이름"]).strip(),
                                성격_판정=str(row["성격 판정"]).strip(),
                                근거=str(row["근거 요약"]).strip(),
                                면접_통합데이터=raw_comment,
                                model=selected_model,
                            )
                            lines = [l.strip() for l in result.splitlines() if l.strip()]
                            df_final.at[idx, "최종_성격_판정"] = lines[0] if len(lines) > 0 else ""
                            df_final.at[idx, "성격_키워드"]    = lines[1] if len(lines) > 1 else ""
                            df_final.at[idx, "최종_근거"]      = lines[2] if len(lines) > 2 else (lines[1] if len(lines) > 1 else "")
                        except Exception as e:
                            df_final.at[idx, "최종_근거"] = f"오류: {e}"

                        df_final.at[idx, "면접_평균점수"] = extract_interview_score(raw_comment)
                        progress.progress((_i + 1) / len(targets), text=f"재분석 중... ({_i+1}/{len(targets)})")

                    progress.empty()

                    df_for_stage3 = df_final.copy()
                    df_for_stage3["성격 판정"]   = df_for_stage3["최종_성격_판정"]
                    df_for_stage3["근거 요약"]   = df_for_stage3["최종_근거"]
                    df_for_stage3["성격_키워드"] = df_for_stage3["성격_키워드"]
                    st.session_state["df_personality"] = df_for_stage3.drop(
                        columns=["최종_성격_판정", "최종_근거", "면접_통합데이터"], errors="ignore"
                    )
                    st.session_state["df_merged"] = df_final
                    st.success("재분석 완료! STEP 3 자리배치에 자동으로 반영돼요.")
                    st.dataframe(df_final, use_container_width=True)

                st.divider()
                excel_bytes = to_final_excel_bytes(st.session_state["df_merged"])
                st.download_button(
                    label="최종 분석 파일 엑셀 다운로드",
                    data=excel_bytes,
                    file_name="최종_분석_결과.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )


# ── 3단계: 성격 조정 ─────────────────────────────────────────
with tab3:
    st.header("✏️ STEP 3 · 성격 조정")
    st.caption("분析 결과를 보고, AI 판단이 틀렸다면 여기서 직접 바꿔주세요. 수정 후 STEP 4로 이동하세요.")

    if "df_personality" not in st.session_state:
        st.warning("먼저 **STEP 2**에서 자소서 분析을 완료해 주세요!")
    else:
        _df_adj_src = st.session_state["df_personality"].copy()

        st.info(f"총 **{len(_df_adj_src)}명** — 이름 옆 토글로 외향/내향을 바꾼 뒤 **확정** 버튼을 눌러주세요.")

        # st.data_editor로 간편 수정
        _display_cols = [c for c in ["이름", "학과", "학번", "성격 판정", "근거 요약", "성격_키워드", "면접_평균점수"]
                         if c in _df_adj_src.columns]
        _edited_df = st.data_editor(
            _df_adj_src[_display_cols].copy(),
            column_config={
                "이름":       st.column_config.TextColumn("이름", disabled=True, width="small"),
                "학과":       st.column_config.TextColumn("학과", disabled=True, width="small"),
                "학번":       st.column_config.TextColumn("학번", disabled=True, width="small"),
                "성격 판정":  st.column_config.SelectboxColumn(
                    "성격 판정",
                    options=["외향형", "내향형"],
                    required=True,
                    width="small",
                ),
                "근거 요약":       st.column_config.TextColumn("근거", disabled=True, width="medium"),
                "성격_키워드":     st.column_config.TextColumn("키워드", disabled=True, width="medium"),
                "면접_평균점수":   st.column_config.TextColumn("면접점수", disabled=True, width="small"),
            },
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            key="ei_editor",
        )

        st.divider()
        _col_a, _col_b = st.columns([1, 3])
        if _col_a.button("✅ 조정 확정 → STEP 4", type="primary", key="btn_confirm_adjust"):
            # 수정된 성격 판정을 df_personality에 반영
            _df_out = _df_adj_src.copy()
            _df_out["성격 판정"] = _edited_df["성격 판정"].values
            st.session_state["df_personality"] = _df_out
            st.session_state["df_ei_adjusted"] = True
            _e = (_df_out["성격 판정"] == "외향형").sum()
            _i = (_df_out["성격 판정"] == "내향형").sum()
            st.success(f"확정 완료! 외향형 {_e}명 / 내향형 {_i}명 → STEP 4 탭으로 이동하세요.")
        _col_b.caption("수정 없이 그냥 넘어가도 괜찮아요. STEP 4에서 분析 결과 그대로 배치돼요.")


# ── 4단계: 자리배치 ───────────────────────────────────────────
with tab4:
    st.header("🪑 STEP 4 · 자리배치")
    st.caption("설정을 맞추고 버튼을 누르면 자리가 자동으로 배치돼요!")

    col_left, col_right = st.columns([1, 2])

    with col_left:
        st.subheader("⚙️ 배치 설정")

        num_people = st.slider("테이블당 최대 인원수", min_value=2, max_value=8, value=4, step=1)

        st.divider()

        personality = st.radio(
            "성격 기준",
            options=["혼합 배치 (외향·내향 섞기)", "유사 배치 (비슷한 성격끼리)", "무작위"],
            index=0,
            help="외향(E)·내향(I) 혼합은 모든 모드에서 항상 적용돼요.",
        )

        st.divider()

        student_id_policy = st.radio(
            "같은 학번 처리 방식",
            options=["학번 무관", "동일 학번 분리", "동일 학번 우선 배치"],
            index=0,
        )

    with col_right:
        st.subheader("배치 결과")

        has_data = "df_personality" in st.session_state
        if not has_data:
            st.info("먼저 STEP 2에서 자소서 성격 분석을 완료해 주세요!")

        if st.button("🪑 자리 배치하기!", type="primary", disabled=not has_data):
            df_src = st.session_state["df_personality"].copy()

            # 취소자 한 번 더 필터 (혹시 분석 이후 추가된 경우)
            _cancel_set3 = {
                n.strip()
                for n in st.session_state.get("cancellation_input", "").splitlines()
                if n.strip()
            }
            if _cancel_set3:
                _keep3 = ~df_src["이름"].astype(str).str.strip().isin(_cancel_set3)
                df_src = df_src[_keep3].reset_index(drop=True)

            late_names = [
                n.strip()
                for n in st.session_state.get("late_arrivals_input", "").splitlines()
                if n.strip()
            ]
            # 임원진 파싱: "이름 외향/내향" 형식. 성격 미기재 시 외향형 기본값
            officer_entries: list[tuple[str, str]] = []
            for _line in st.session_state.get("officer_input", "").splitlines():
                _line = _line.strip()
                if not _line:
                    continue
                if "내향" in _line:
                    _ei = "내향형"
                else:
                    _ei = "외향형"
                # 이름: 성격 키워드(외향/내향/형) 제거 후 앞부분
                _name = _line.replace("외향형", "").replace("내향형", "").replace("외향", "").replace("내향", "").strip()
                if _name:
                    officer_entries.append((_name, _ei))

            extra_rows: list[dict] = []

            for _oname, _oei in officer_entries:
                extra_rows.append({
                    "이름":        _oname,
                    "학과":        "미상",
                    "학번":        "",
                    "성격 판정":   _oei,
                    "근거 요약":   "임원/기존부원",
                    "성격_키워드": "",
                    "임원":        True,
                })

            _officer_e_cnt = sum(1 for _, _ei in officer_entries if _ei == "외향형")
            e_cnt = df_src["성격 판정"].str.contains("외향", na=False).sum() + _officer_e_cnt
            i_cnt = len(df_src) - (df_src["성격 판정"].str.contains("외향", na=False).sum())
            for name in late_names:
                ei_val = "외향형" if e_cnt <= i_cnt else "내향형"
                if ei_val == "외향형":
                    e_cnt += 1
                else:
                    i_cnt += 1
                extra_rows.append({
                    "이름":        name,
                    "학과":        "미상",
                    "학번":        "",
                    "성격 판정":   ei_val,
                    "근거 요약":   "늦참자",
                    "성격_키워드": "",
                    "늦참자":      True,
                })

            if extra_rows:
                df_src = pd.concat([df_src, pd.DataFrame(extra_rows)], ignore_index=True)
            # NaN → False 로 채워야 bool 변환 시 오탐 없음
            df_src["임원"]   = df_src.get("임원",   pd.Series(False, index=df_src.index)).fillna(False)
            df_src["늦참자"] = df_src.get("늦참자", pd.Series(False, index=df_src.index)).fillna(False)

            # 성격 판정 값 정규화 (LLM 출력이 "외향형입니다" 등일 수 있어서)
            df_src["성격 판정"] = df_src["성격 판정"].apply(
                lambda x: "외향형" if "외향" in str(x) else "내향형"
            )
            with st.spinner("자리 배치 중..."):
                df_seated = assign_seats(
                    df_src, num_people, personality, student_id_policy,
                    model=selected_model,
                )

            st.session_state["df_seated"] = df_seated
            total    = len(df_seated)
            n_tables = df_seated["테이블_번호"].max()
            late_cnt = int(df_seated["늦참자"].fillna(False).sum()) if "늦참자" in df_seated.columns else 0
            off_cnt  = int(df_seated["임원"].fillna(False).sum())   if "임원"   in df_seated.columns else 0
            extra_msg = []
            if off_cnt:  extra_msg.append(f"임원/기존부원 {off_cnt}명")
            if late_cnt: extra_msg.append(f"늦참자 {late_cnt}명")
            extra_str = f" ({', '.join(extra_msg)} 포함)" if extra_msg else ""
            st.success(f"배치 완료! 총 {total}명을 {n_tables}개 테이블에 나눴어요.{extra_str}")

        if "df_seated" in st.session_state:
            df_seated = st.session_state["df_seated"]

            # ── 테이블 카드 뷰 ────────────────────────────────
            st.divider()
            st.markdown("#### 테이블별 구성")

            COLS_PER_ROW = 3
            table_nums = sorted(df_seated["테이블_번호"].unique())

            for row_start in range(0, len(table_nums), COLS_PER_ROW):
                card_cols = st.columns(COLS_PER_ROW)
                for col_i, t_num in enumerate(table_nums[row_start: row_start + COLS_PER_ROW]):
                    grp   = df_seated[df_seated["테이블_번호"] == t_num]
                    e_cnt = (grp["EI"] == "E").sum()
                    i_cnt = (grp["EI"] == "I").sum()
                    with card_cols[col_i]:
                        st.markdown(f"**테이블 {t_num}** &nbsp; `E:{e_cnt} I:{i_cnt}`")
                        for _, r in grp.iterrows():
                            gender_icon = {"남": "♂", "여": "♀"}.get(r["성별"], "?")
                            ei_color = "#d4a017" if r["EI"] == "E" else "#27ae60"
                            ei_tag = (
                                f'<span style="color:{ei_color};font-weight:bold">'
                                f'{r["EI"]}</span>'
                            )
                            tooltip = str(r.get("성격_키워드", "")).strip()
                            if not tooltip or tooltip in ("nan", "None"):
                                tooltip = ""
                            is_late    = r.get("늦참자") is True
                            is_officer = r.get("임원") is True
                            late_badge = ' <span style="color:#e67e22;font-size:0.75em">⏰늦참</span>' if is_late    else ""
                            off_badge  = ' <span style="color:#8e44ad;font-size:0.75em">👑임원</span>' if is_officer else ""
                            name_html = (
                                f'<span title="{tooltip}" style="cursor:help;'
                                f'border-bottom:1px dotted #888">{r["이름"]}</span>'
                            )
                            dept_str = r["학과"] if str(r["학과"]) not in ("미상", "nan", "") else "?"
                            year_str = r["학번_연도"] if str(r["학번_연도"]) not in ("미상", "nan", "") else "?"
                            st.markdown(
                                f"- {gender_icon} {name_html} ({dept_str} / {year_str}학번) {ei_tag}{late_badge}{off_badge}",
                                unsafe_allow_html=True,
                            )

            # ── 전체 데이터프레임 ─────────────────────────────
            st.divider()
            st.markdown("#### 전체 목록")
            display_cols = ["테이블_번호", "이름", "학과", "학번", "성별", "EI", "성격 판정"]
            st.dataframe(df_seated[display_cols], use_container_width=True)

            # ── 엑셀 다운로드 ────────────────────────────────
            st.divider()
            excel_bytes = to_excel_bytes(df_seated[display_cols + ["근거 요약"]])
            st.download_button(
                label="엑셀로 다운로드",
                data=excel_bytes,
                file_name="자리배치_결과.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
